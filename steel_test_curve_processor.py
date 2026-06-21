"""
Process RAW test data of structural steel tests.

Put this script in the same folder as the RAW .xlsx files, then run:
    python steel-test-curve-processor.py

Required packages:
    pip install pandas openpyxl numpy

Main outputs:
    processed_test_curves.xlsx

Author: Maxiang He
"""

from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# USER INPUT AREA
# =============================================================================

# 1-based column numbers in your RAW Excel files.
# Example: X_COL = 5 means the 5th column is used as x.
X_COL = 5 # strain or displacement
Y_COL = 12 # load

# Sheet name or index. Use 0 for the first sheet. If all files only have one sheet,
# keep SHEET = 0.
SHEET = 0

# Output file name.
OUTPUT_XLSX = "processed_test_curves.xlsx"

# -----------------------------------------------------------------------------
# DENOISING / DOWNSAMPLING PARAMETERS YOU WILL MOST OFTEN TUNE
# -----------------------------------------------------------------------------
# Use larger values if you want smoother and sparser curves.
# Suggested trial values:
#   0.7 = weaker denoising, more points
#   1.0 = default
#   1.5 = stronger denoising, fewer points
#   2.0 = very strong denoising, much fewer points
ELASTIC_DENOISE_STRENGTH = 1
POST_ELASTIC_DENOISE_STRENGTH = 1

# Approximate final number of points per curve before strength factors are applied.
# The actual final number may be slightly higher because peak and fracture/drop
# points are preserved.
BASE_TARGET_POINTS_PER_CURVE = 260

# Maximum number of points kept in the initial elastic range.
# Smaller value = sparser and cleaner elastic range.
BASE_ELASTIC_MAX_POINTS = 25

# Smoothing window sizes. Larger values = smoother curves.
# Elastic smoothing is used before elastic linear reconstruction; post-elastic
# smoothing is used for ascending/descending branches after the linear elastic part.
BASE_ELASTIC_SMOOTHING_WINDOW_FRACTION = 0.045
BASE_POST_ELASTIC_SMOOTHING_WINDOW_FRACTION = 0.020
BASE_MIN_SMOOTHING_WINDOW = 5
BASE_MAX_SMOOTHING_WINDOW = 61

# Automatic data-start detection.
# The script searches for the first row where x and y are numeric for several consecutive rows.
MIN_CONSECUTIVE_NUMERIC_ROWS = 5

# Decide whether x is strain or displacement.
# If this fraction of abs(x) is < 1, x is treated as strain; otherwise displacement.
STRAIN_FRACTION_LIMIT = 0.80

# Initial slip correction settings.
# The script detects the first true straight elastic segment automatically and
# uses it to correct the x-origin. The following y-ratio values are kept only
# for fallback/compatibility.
ENABLE_INITIAL_SLIP_CORRECTION = True
ELASTIC_FIT_Y_MIN_RATIO = 0.20
ELASTIC_FIT_Y_MAX_RATIO = 0.60
MIN_POINTS_FOR_ELASTIC_FIT = 20

# Elastic range denoising.
# Only the first detected straight elastic segment is reconstructed as a pure
# straight line. Later nonlinear ascending response is NOT treated as elastic.
ENABLE_ELASTIC_LINEARISATION = True
ELASTIC_LINEARISE_Y_MAX_RATIO = 0.65
ELASTIC_LINEAR_BLEND_ZONE_FRACTION = 0.20

# The script identifies elastic range as ONLY the first straight rising segment.
# It tests progressively longer initial segments and keeps the longest segment
# that is still sufficiently linear. Increase INITIAL_ELASTIC_PREFIX_MIN_R2 if
# you want a shorter/purer elastic range; decrease it if the raw elastic data are noisy.
INITIAL_ELASTIC_DETECTION_MIN_Y_RATIO = 0.03
INITIAL_ELASTIC_DETECTION_MAX_Y_RATIO = 0.35
INITIAL_ELASTIC_END_SEARCH_MAX_Y_RATIO = 0.78
INITIAL_ELASTIC_PREFIX_MIN_R2 = 0.990
INITIAL_ELASTIC_MIN_POINTS = 80
INITIAL_ELASTIC_CANDIDATE_STEP = 5
INITIAL_ELASTIC_SLOPE_DROP_CONFIRMATION = 0.65
INITIAL_ELASTIC_DEPARTURE_TOL_RATIO_OF_PEAK = 0.030
INITIAL_ELASTIC_DEPARTURE_TOL_RATIO_OF_LINE = 0.120
INITIAL_ELASTIC_DEPARTURE_CONSECUTIVE_POINTS = 8


# Smoothing switch.
# Ascending branch: sudden local noise is smoothed.
# Descending branch: large force drops are preserved; smoothing is only applied between drops.
ENABLE_SMOOTHING = True

# Descending branch sudden-drop detection.
# A drop is preserved if it is either larger than FORCE_DROP_ABSOLUTE_KN, or larger than
# FORCE_DROP_RATIO_OF_PEAK * peak force, and also stands out relative to local noise.
FORCE_DROP_ABSOLUTE_KN = 5.0
FORCE_DROP_RATIO_OF_PEAK = 0.015
FORCE_DROP_MAD_MULTIPLIER = 8.0

# Remove points before the corrected x-origin. This normally removes initial seating/slip.
DROP_NEGATIVE_CORRECTED_X = True

# Structural tests are normally monotonic in strain/displacement. Enforcing this
# removes small backward jumps in noisy x readings, especially near the elastic/nonlinear transition.
ENFORCE_MONOTONIC_X = True

# Files matching these names will not be processed.
IGNORE_FILE_PATTERNS = [
    OUTPUT_XLSX,
    "processed*.xlsx",
    "~$*.xlsx",
]

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================


def is_ignored_file(path: Path) -> bool:
    for pat in IGNORE_FILE_PATTERNS:
        if pat.startswith("~$") and path.name.startswith("~$"):
            return True
        if path.match(pat):
            return True
    return False


def to_number(value) -> float:
    """Convert cell value to float. Returns NaN if conversion fails."""
    if value is None:
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    text = str(value).strip()
    if text == "":
        return np.nan
    # Remove common non-numeric characters, while keeping signs, decimal points and exponents.
    text = text.replace(",", "")
    match = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", text)
    if not match:
        return np.nan
    try:
        return float(match.group(0))
    except ValueError:
        return np.nan


def read_raw_xy(path: Path, x_col: int, y_col: int, sheet=0) -> Tuple[np.ndarray, np.ndarray, int]:
    """Read selected x/y columns and automatically locate the first data row."""
    # Read without header so header/unit rows remain available for automatic detection.
    df = pd.read_excel(path, sheet_name=sheet, header=None, engine="openpyxl")

    if x_col < 1 or y_col < 1:
        raise ValueError("X_COL and Y_COL must be 1-based positive column numbers.")
    if x_col > df.shape[1] or y_col > df.shape[1]:
        raise ValueError(
            f"Selected columns exceed available columns. File has {df.shape[1]} columns."
        )

    x_raw = df.iloc[:, x_col - 1].map(to_number).to_numpy(dtype=float)
    y_raw = df.iloc[:, y_col - 1].map(to_number).to_numpy(dtype=float)
    valid = np.isfinite(x_raw) & np.isfinite(y_raw)

    start_idx: Optional[int] = None
    count = 0
    for i, ok in enumerate(valid):
        if ok:
            count += 1
            if count >= MIN_CONSECUTIVE_NUMERIC_ROWS:
                start_idx = i - MIN_CONSECUTIVE_NUMERIC_ROWS + 1
                break
        else:
            count = 0

    if start_idx is None:
        raise ValueError("Cannot find consecutive numeric rows in the selected x/y columns.")

    x = x_raw[start_idx:]
    y = y_raw[start_idx:]
    mask = np.isfinite(x) & np.isfinite(y)
    x = x[mask]
    y = y[mask]

    # Remove fully repeated adjacent points if they exist.
    if len(x) > 1:
        keep = np.ones(len(x), dtype=bool)
        keep[1:] = ~((np.diff(x) == 0) & (np.diff(y) == 0))
        x, y = x[keep], y[keep]

    if len(x) < 10:
        raise ValueError("Too few valid data points after cleaning.")

    return x, y, start_idx + 1  # return 1-based Excel row number


def classify_x_type(x_abs: np.ndarray) -> str:
    """Classify x as strain or displacement based on most values being below/above 1."""
    finite = x_abs[np.isfinite(x_abs)]
    if len(finite) == 0:
        return "unknown"
    fraction_less_than_1 = np.mean(finite < 1.0)
    if fraction_less_than_1 >= STRAIN_FRACTION_LIMIT:
        return "strain"
    return "displacement_mm"


def clamp_strength(value: float) -> float:
    """Avoid zero/negative or unrealistically large denoising factors."""
    try:
        value = float(value)
    except Exception:
        return 1.0
    return min(max(value, 0.2), 5.0)


def effective_total_target_points() -> int:
    """Larger POST_ELASTIC_DENOISE_STRENGTH gives fewer overall points."""
    strength = clamp_strength(POST_ELASTIC_DENOISE_STRENGTH)
    return max(80, int(round(BASE_TARGET_POINTS_PER_CURVE / strength)))


def effective_elastic_max_points() -> int:
    """Larger ELASTIC_DENOISE_STRENGTH gives fewer elastic points."""
    strength = clamp_strength(ELASTIC_DENOISE_STRENGTH)
    return max(6, int(round(BASE_ELASTIC_MAX_POINTS / strength)))


def odd_window(n: int, base_fraction: float, strength: float) -> int:
    strength = clamp_strength(strength)
    w = int(round(n * base_fraction * strength))
    min_w = BASE_MIN_SMOOTHING_WINDOW
    max_w = int(round(BASE_MAX_SMOOTHING_WINDOW * strength))
    max_w = max(min_w, max_w)
    w = max(min_w, min(max_w, w))
    if w % 2 == 0:
        w += 1
    if w >= n:
        w = n - 1 if (n - 1) % 2 == 1 else n - 2
    return max(3, w)


def smooth_series(
    y: np.ndarray,
    base_fraction: float = BASE_POST_ELASTIC_SMOOTHING_WINDOW_FRACTION,
    strength: float = POST_ELASTIC_DENOISE_STRENGTH,
) -> np.ndarray:
    """Robust smoothing using rolling median followed by rolling mean."""
    if not ENABLE_SMOOTHING or len(y) < BASE_MIN_SMOOTHING_WINDOW:
        return y.copy()
    w = odd_window(len(y), base_fraction, strength)
    s = pd.Series(y)
    y_med = s.rolling(window=w, center=True, min_periods=1).median()
    y_smooth = y_med.rolling(window=w, center=True, min_periods=1).mean()
    return y_smooth.to_numpy(dtype=float)


def fit_line_with_stats(xf: np.ndarray, yf: np.ndarray) -> Tuple[float, float, float]:
    """Fit y = k*x + c and return k, c and R2."""
    valid = np.isfinite(xf) & np.isfinite(yf)
    xf, yf = xf[valid], yf[valid]
    if len(xf) < 3 or np.nanstd(xf) <= 0:
        return np.nan, np.nan, np.nan
    k, c = np.polyfit(xf, yf, 1)
    if not np.isfinite(k) or not np.isfinite(c):
        return np.nan, np.nan, np.nan
    y_pred = k * xf + c
    ss_res = float(np.sum((yf - y_pred) ** 2))
    ss_tot = float(np.sum((yf - np.mean(yf)) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
    return float(k), float(c), float(r2)


def robust_refit_line(xf: np.ndarray, yf: np.ndarray, min_points: int) -> Tuple[float, float, float]:
    """Fit a line, remove strong residual outliers, then refit."""
    k, c, r2 = fit_line_with_stats(xf, yf)
    if not np.isfinite(k) or abs(k) < 1e-12:
        return np.nan, np.nan, np.nan

    residual = yf - (k * xf + c)
    mad = np.median(np.abs(residual - np.median(residual)))
    if mad > 0:
        inliers = np.abs(residual - np.median(residual)) <= 3.0 * 1.4826 * mad
        if np.count_nonzero(inliers) >= min_points:
            k, c, r2 = fit_line_with_stats(xf[inliers], yf[inliers])
    return k, c, r2


def detect_initial_straight_elastic_segment(
    x: np.ndarray, y: np.ndarray, y_peak: float
) -> Dict[str, float]:
    """
    Detect ONLY the first straight rising elastic segment.

    Method used here:
    1. Smooth y only for trend detection.
    2. Ignore the very first seating/slip region when fitting.
    3. Progressively fit longer prefixes of the initial rising branch.
    4. Choose the longest prefix that is still sufficiently linear.
    5. Optionally trim the end if the following local tangent stiffness has already
       dropped strongly, which indicates the start of yielding/nonlinear response.

    This is deliberately different from treating the full ascending branch as elastic.
    """
    out: Dict[str, float] = {
        "ok": 0,
        "k": np.nan,
        "c": np.nan,
        "r2": np.nan,
        "x_intercept": np.nan,
        "fit_start_idx": 0,
        "fit_end_idx": 0,
        "elastic_end_idx": 0,
    }

    if len(x) < MIN_POINTS_FOR_ELASTIC_FIT or y_peak <= 0:
        return out

    peak_idx = int(np.nanargmax(y))
    if peak_idx < MIN_POINTS_FOR_ELASTIC_FIT:
        return out

    asc_x = np.asarray(x[: peak_idx + 1], dtype=float)
    asc_y = np.asarray(y[: peak_idx + 1], dtype=float)

    # Sort out small backward jumps in x before fitting/detection.
    if ENFORCE_MONOTONIC_X:
        asc_x = np.maximum.accumulate(asc_x)

    # Smooth only for detection; raw data are still used for final fitting.
    y_det = smooth_series(
        asc_y,
        base_fraction=BASE_ELASTIC_SMOOTHING_WINDOW_FRACTION,
        strength=max(1.0, ELASTIC_DENOISE_STRENGTH),
    )

    # Skip the first seating/slip part when fitting, but keep it in the final elastic range.
    fit_start_candidates = np.where(y_det >= INITIAL_ELASTIC_DETECTION_MIN_Y_RATIO * y_peak)[0]
    fit_start = int(fit_start_candidates[0]) if len(fit_start_candidates) else 0
    fit_start = min(fit_start, max(0, peak_idx - MIN_POINTS_FOR_ELASTIC_FIT))

    max_end_candidates = np.where(y_det <= INITIAL_ELASTIC_END_SEARCH_MAX_Y_RATIO * y_peak)[0]
    max_end = int(max_end_candidates[-1]) if len(max_end_candidates) else int(0.60 * peak_idx)
    max_end = min(max_end, peak_idx - 3)

    min_end = max(fit_start + INITIAL_ELASTIC_MIN_POINTS, fit_start + MIN_POINTS_FOR_ELASTIC_FIT)
    if min_end >= max_end:
        min_end = min(max_end, fit_start + MIN_POINTS_FOR_ELASTIC_FIT)
    if max_end - fit_start < MIN_POINTS_FOR_ELASTIC_FIT:
        return out

    best = None
    last_good = None
    # Choose the longest prefix with acceptable R2. This matches the engineering
    # definition of elastic range: the whole initial segment before sustained nonlinearity.
    for end_idx in range(min_end, max_end + 1, max(1, INITIAL_ELASTIC_CANDIDATE_STEP)):
        xf = asc_x[fit_start : end_idx + 1]
        yf = asc_y[fit_start : end_idx + 1]
        if len(xf) < MIN_POINTS_FOR_ELASTIC_FIT or np.nanstd(xf) <= 0:
            continue
        k, c, r2 = robust_refit_line(xf, yf, max(8, MIN_POINTS_FOR_ELASTIC_FIT // 2))
        if not np.isfinite(k) or k <= 0 or not np.isfinite(r2):
            continue

        # Additional check: the measured curve should not have persistently departed
        # from the fitted line within this candidate prefix.
        y_line = k * asc_x[fit_start : end_idx + 1] + c
        residual = np.abs(y_det[fit_start : end_idx + 1] - y_line)
        tol = np.maximum(
            INITIAL_ELASTIC_DEPARTURE_TOL_RATIO_OF_PEAK * y_peak,
            INITIAL_ELASTIC_DEPARTURE_TOL_RATIO_OF_LINE * np.maximum(y_line, 0.05 * y_peak),
        )
        good_departure = np.mean(residual <= tol) >= 0.92

        if best is None or r2 > best[0]:
            best = (r2, end_idx, k, c)
        if r2 >= INITIAL_ELASTIC_PREFIX_MIN_R2 and good_departure:
            last_good = (r2, end_idx, k, c)

    chosen = last_good if last_good is not None else best
    if chosen is None:
        return out

    r2, elastic_end, k, c = chosen

    # Confirm with local tangent stiffness after the candidate end. If the slope
    # has already dropped greatly, trim a little earlier. If not, keep the chosen end.
    try:
        dx = np.diff(asc_x)
        dy = np.diff(y_det)
        valid = dx > np.nanmedian(dx[dx > 0]) * 0.05 if np.any(dx > 0) else dx > 0
        local_slope = np.full_like(dy, np.nan, dtype=float)
        local_slope[valid] = dy[valid] / dx[valid]
        w = max(9, min(41, int(round(0.03 * len(asc_x)))))
        if w % 2 == 0:
            w += 1
        slope_s = pd.Series(local_slope).rolling(window=w, center=True, min_periods=1).median().to_numpy(float)
        base_slope = np.nanmedian(slope_s[fit_start : max(fit_start + 5, min(elastic_end, fit_start + 80))])
        if np.isfinite(base_slope) and base_slope > 0:
            post_start = min(elastic_end + 1, len(slope_s) - 1)
            post_end = min(len(slope_s), elastic_end + 1 + max(10, w))
            post_slope = np.nanmedian(slope_s[post_start:post_end]) if post_start < post_end else np.nan
            # This is only a confirmation/trim, not the primary criterion.
            if np.isfinite(post_slope) and post_slope < INITIAL_ELASTIC_SLOPE_DROP_CONFIRMATION * base_slope:
                elastic_end = max(fit_start + MIN_POINTS_FOR_ELASTIC_FIT, elastic_end - w // 3)
                k, c, r2 = robust_refit_line(asc_x[fit_start : elastic_end + 1], asc_y[fit_start : elastic_end + 1], max(8, MIN_POINTS_FOR_ELASTIC_FIT // 2))
    except Exception:
        pass

    x_intercept = -c / k if np.isfinite(k) and abs(k) > 1e-12 else np.nan
    if not np.isfinite(x_intercept):
        return out

    out.update(
        {
            "ok": 1,
            "k": float(k),
            "c": float(c),
            "r2": float(r2),
            "x_intercept": float(x_intercept),
            "fit_start_idx": int(fit_start),
            "fit_end_idx": int(elastic_end),
            "elastic_end_idx": int(elastic_end),
        }
    )
    return out

def robust_elastic_fit(x: np.ndarray, y: np.ndarray, peak_idx: int, y_peak: float) -> Tuple[float, float, float]:
    """Fit the first detected straight elastic segment and return y = k*x + c and R2."""
    det = detect_initial_straight_elastic_segment(x, y, y_peak)
    if det.get("ok", 0) != 1:
        return np.nan, np.nan, np.nan
    return float(det["k"]), float(det["c"]), float(det["r2"])


def linearise_elastic_branch(
    x_asc: np.ndarray, y_asc_smoothed: np.ndarray, y_asc_raw: np.ndarray, y_peak: float
) -> Tuple[np.ndarray, int, Dict[str, float]]:
    """
    Replace only the initial straight elastic range with a pure straight line.

    The elastic range is detected from the first stable straight part of the curve,
    not from the full ascending branch. This is intended to remove initial slip /
    seating distortion while leaving the later nonlinear ascending branch untouched.
    """
    info = {
        "elastic_linearisation_applied": 0,
        "elastic_linearised_points": 0,
        "elastic_linear_fit_slope": np.nan,
        "elastic_linear_fit_intercept": np.nan,
        "elastic_linear_fit_r2": np.nan,
    }
    if not ENABLE_ELASTIC_LINEARISATION or len(x_asc) < MIN_POINTS_FOR_ELASTIC_FIT or y_peak <= 0:
        return y_asc_smoothed.copy(), 0, info

    det = detect_initial_straight_elastic_segment(x_asc, y_asc_raw, y_peak)
    if det.get("ok", 0) != 1:
        return y_asc_smoothed.copy(), 0, info

    k = float(det["k"])
    c = float(det["c"])
    r2 = float(det["r2"])
    elastic_end = int(det["elastic_end_idx"])
    if not np.isfinite(k) or k <= 0 or elastic_end < MIN_POINTS_FOR_ELASTIC_FIT:
        return y_asc_smoothed.copy(), 0, info

    y_out = y_asc_smoothed.copy()

    # Replace the true initial elastic range with one pure straight line.
    # To remove initial slip completely, the final line is forced through the
    # origin and through the detected elastic-end point. This avoids carrying
    # over the slightly crooked first few measured points.
    y_anchor = float(y_asc_smoothed[elastic_end])
    if not np.isfinite(y_anchor) or y_anchor <= 0:
        y_anchor = float(y_asc_raw[elastic_end])
    y_anchor = max(0.0, y_anchor)
    y_out[: elastic_end + 1] = np.linspace(0.0, y_anchor, elastic_end + 1)

    # Force the processed curve to start from the origin.
    if len(y_out) > 0:
        y_out[0] = 0.0

    info.update(
        {
            "elastic_linearisation_applied": 1,
            "elastic_linearised_points": int(elastic_end + 1),
            "elastic_linear_fit_slope": float(k),
            "elastic_linear_fit_intercept": float(c),
            "elastic_linear_fit_r2": float(r2),
            "elastic_fit_start_idx": int(det.get("fit_start_idx", 0)),
            "elastic_fit_end_idx": int(det.get("fit_end_idx", elastic_end)),
        }
    )
    return y_out, elastic_end, info


def linear_fit_initial_slip(x: np.ndarray, y: np.ndarray) -> Tuple[np.ndarray, Dict[str, float]]:
    """
    Correct initial slip/seating using the first stable straight elastic segment.

    The detected line is y = k*x + c. Its x-intercept is x0 = -c/k.
    The corrected coordinate is x_corr = x - x0, so the real elastic line
    is shifted to pass through the origin.
    """
    info: Dict[str, float] = {
        "slip_correction_applied": 0,
        "x_intercept_removed": 0.0,
        "elastic_slope": np.nan,
        "elastic_fit_r2": np.nan,
    }

    if not ENABLE_INITIAL_SLIP_CORRECTION or len(x) < MIN_POINTS_FOR_ELASTIC_FIT:
        return x.copy(), info

    peak_idx = int(np.nanargmax(y))
    if peak_idx < MIN_POINTS_FOR_ELASTIC_FIT:
        return x.copy(), info

    y_peak = float(np.nanmax(y))
    if y_peak <= 0:
        return x.copy(), info

    det = detect_initial_straight_elastic_segment(x, y, y_peak)
    if det.get("ok", 0) != 1:
        return x.copy(), info

    k = float(det["k"])
    x0 = float(det["x_intercept"])
    r2 = float(det["r2"])

    # Avoid unreasonable shifts; seating/slip should be a small part of x at peak.
    x_peak = x[peak_idx]
    if not np.isfinite(x0) or not np.isfinite(x_peak) or not np.isfinite(k):
        return x.copy(), info
    if abs(x0) > 0.5 * max(abs(x_peak), 1e-12):
        return x.copy(), info

    x_corr = x - x0
    info.update(
        {
            "slip_correction_applied": 1,
            "x_intercept_removed": float(x0),
            "elastic_slope": float(k),
            "elastic_fit_r2": float(r2),
        }
    )
    return x_corr, info


def detect_descending_drops(y_desc: np.ndarray, y_peak: float) -> np.ndarray:
    """Return indices in descending-branch local coordinates around sudden force drops."""
    if len(y_desc) < 5 or y_peak <= 0:
        return np.array([], dtype=int)
    dy = np.diff(y_desc)
    negative_drop = -dy  # positive when y suddenly drops
    mad = np.median(np.abs(negative_drop - np.median(negative_drop)))
    robust_limit = FORCE_DROP_MAD_MULTIPLIER * 1.4826 * mad if mad > 0 else 0.0
    absolute_limit = max(FORCE_DROP_ABSOLUTE_KN, FORCE_DROP_RATIO_OF_PEAK * y_peak)
    threshold = max(absolute_limit, robust_limit)
    jump_before = np.where(negative_drop > threshold)[0]

    keep = []
    for j in jump_before:
        # Preserve points immediately before and after the drop.
        keep.extend([j, j + 1])
    keep = sorted(set(i for i in keep if 0 <= i < len(y_desc)))
    return np.array(keep, dtype=int)


def lttb_downsample(x: np.ndarray, y: np.ndarray, n_out: int) -> np.ndarray:
    """
    Largest-Triangle-Three-Buckets downsampling.
    Returns selected indices preserving the visual shape of the curve.
    """
    n = len(x)
    if n_out >= n or n_out <= 0:
        return np.arange(n)
    if n_out < 3:
        return np.array([0, n - 1], dtype=int)

    sampled = [0]
    bucket_size = (n - 2) / (n_out - 2)
    a = 0

    for i in range(n_out - 2):
        start = int(math.floor((i + 0) * bucket_size)) + 1
        end = int(math.floor((i + 1) * bucket_size)) + 1
        end = min(end, n - 1)

        next_start = int(math.floor((i + 1) * bucket_size)) + 1
        next_end = int(math.floor((i + 2) * bucket_size)) + 1
        next_end = min(next_end, n)

        if next_start >= next_end:
            avg_x = x[-1]
            avg_y = y[-1]
        else:
            avg_x = np.mean(x[next_start:next_end])
            avg_y = np.mean(y[next_start:next_end])

        range_x = x[start:end]
        range_y = y[start:end]
        if len(range_x) == 0:
            continue

        area = np.abs(
            (x[a] - avg_x) * (range_y - y[a]) - (x[a] - range_x) * (avg_y - y[a])
        )
        idx = start + int(np.argmax(area))
        sampled.append(idx)
        a = idx

    sampled.append(n - 1)
    return np.array(sorted(set(sampled)), dtype=int)


def process_curve(x_raw: np.ndarray, y_raw: np.ndarray) -> Tuple[np.ndarray, np.ndarray, Dict[str, float | str]]:
    """Apply absolute value, classify x, correct slip, smooth, and downsample."""
    x = np.abs(np.asarray(x_raw, dtype=float))
    y = np.abs(np.asarray(y_raw, dtype=float))

    x_type = classify_x_type(x)

    if x_type == "displacement_mm":
        x = x - x[0]

    x, slip_info = linear_fit_initial_slip(x, y)

    if DROP_NEGATIVE_CORRECTED_X:
        keep = x >= 0
        # Always retain at least some data even if correction is too aggressive.
        if np.count_nonzero(keep) >= 10:
            x, y = x[keep], y[keep]

    # Re-zero after any removal. For displacement this is essential; for strain it is also convenient
    # for plotting a corrected test curve from origin.
    x = x - x[0]
    x[x < 0] = 0.0

    peak_idx = int(np.nanargmax(y))
    y_peak = float(np.nanmax(y))

    asc_x, asc_y = x[: peak_idx + 1], y[: peak_idx + 1]
    desc_x, desc_y = x[peak_idx:], y[peak_idx:]

    # Smooth ascending branch fully, then strongly linearise the elastic range.
    # Use the stronger elastic smoothing first because the initial branch is usually
    # the noisiest part of structural test curves.
    asc_y_s = smooth_series(
        asc_y,
        base_fraction=BASE_ELASTIC_SMOOTHING_WINDOW_FRACTION,
        strength=ELASTIC_DENOISE_STRENGTH,
    )
    elastic_info: Dict[str, float] = {}
    elastic_end_idx = 0
    if len(asc_y_s) > 0:
        asc_y_s, elastic_end_idx, elastic_info = linearise_elastic_branch(
            asc_x, asc_y_s, asc_y, y_peak
        )
        # The early strain/displacement readings can also be noisy. For the elastic
        # region, reconstruct a clean monotonic x-coordinate and a straight y-coordinate.
        # This is only applied before the nonlinear/yielding part.
        if elastic_info.get("elastic_linearisation_applied", 0) == 1 and elastic_end_idx > 2:
            x_end = float(np.nanmax(asc_x[: elastic_end_idx + 1]))
            y_end = float(asc_y_s[elastic_end_idx])
            if np.isfinite(x_end) and np.isfinite(y_end) and x_end > 0 and y_end > 0:
                clean_x = np.linspace(0.0, x_end, elastic_end_idx + 1)
                clean_y = np.linspace(0.0, y_end, elastic_end_idx + 1)
                asc_x[: elastic_end_idx + 1] = clean_x
                asc_y_s[: elastic_end_idx + 1] = clean_y
        # Keep exact peak force value to avoid changing key resistance too much.
        asc_y_s[-1] = asc_y[-1]

    # Smooth descending branch only between sudden force drops.
    if len(desc_y) > 0:
        drop_local = detect_descending_drops(desc_y, y_peak)
        boundaries = sorted(set([0, len(desc_y) - 1] + drop_local.tolist()))
        desc_y_s = desc_y.copy()
        for a, b in zip(boundaries[:-1], boundaries[1:]):
            if b - a + 1 >= BASE_MIN_SMOOTHING_WINDOW:
                desc_y_s[a : b + 1] = smooth_series(
                    desc_y[a : b + 1],
                    base_fraction=BASE_POST_ELASTIC_SMOOTHING_WINDOW_FRACTION,
                    strength=POST_ELASTIC_DENOISE_STRENGTH,
                )
        # Restore exact values at detected drops.
        desc_y_s[drop_local] = desc_y[drop_local]
    else:
        drop_local = np.array([], dtype=int)
        desc_y_s = desc_y

    x_s = np.concatenate([asc_x, desc_x[1:]]) if len(desc_x) > 1 else asc_x
    y_s = np.concatenate([asc_y_s, desc_y_s[1:]]) if len(desc_y_s) > 1 else asc_y_s

    if ENFORCE_MONOTONIC_X and len(x_s) > 1:
        x_s = np.maximum.accumulate(x_s)

    # Downsample each branch, but preserve peak and detected descending drops.
    # The elastic range is almost linear after correction, so keep it deliberately sparse.
    target_points = effective_total_target_points()
    n_asc_target = max(35, int(target_points * 0.52))
    n_desc_target = max(25, target_points - n_asc_target)

    if elastic_end_idx > MIN_POINTS_FOR_ELASTIC_FIT and elastic_end_idx < len(asc_x) - 3:
        n_elastic = min(effective_elastic_max_points(), max(6, int(n_asc_target * 0.22)))
        elastic_idx = lttb_downsample(
            asc_x[: elastic_end_idx + 1],
            asc_y_s[: elastic_end_idx + 1],
            min(n_elastic, elastic_end_idx + 1),
        )

        post_start = elastic_end_idx
        post_x = asc_x[post_start:]
        post_y = asc_y_s[post_start:]
        n_post = max(20, n_asc_target - len(elastic_idx))
        post_idx_local = lttb_downsample(post_x, post_y, min(n_post, len(post_x)))
        post_idx = post_start + post_idx_local
        asc_idx = np.unique(np.concatenate([elastic_idx, post_idx, np.array([0, peak_idx])]))
    else:
        asc_idx = lttb_downsample(asc_x, asc_y_s, min(n_asc_target, len(asc_x)))

    if len(desc_x) > 1:
        desc_idx = lttb_downsample(desc_x, desc_y_s, min(n_desc_target, len(desc_x)))
        desc_idx = np.unique(np.concatenate([desc_idx, drop_local]))
        desc_global_idx = peak_idx + desc_idx
    else:
        desc_global_idx = np.array([peak_idx], dtype=int)

    keep_global = np.unique(np.concatenate([asc_idx, desc_global_idx]))
    keep_global = keep_global[(keep_global >= 0) & (keep_global < len(x_s))]

    x_out = x_s[keep_global]
    y_out = y_s[keep_global]

    # Remove any non-finite values and sort by original sequence, not by x.
    finite = np.isfinite(x_out) & np.isfinite(y_out)
    x_out, y_out = x_out[finite], y_out[finite]

    # Ensure first point is exactly from x = 0.
    if len(x_out) > 0:
        x_out[0] = 0.0

    info: Dict[str, float | str] = {
        "x_type": x_type,
        "raw_points": int(len(x_raw)),
        "processed_points": int(len(x_out)),
        "peak_y_kN_raw_abs": float(y_peak),
        "peak_y_kN_processed": float(np.nanmax(y_out)) if len(y_out) else np.nan,
        "detected_descending_drops": int(len(drop_local) // 2),
    }
    info.update(slip_info)
    info.update(elastic_info)
    return x_out, y_out, info


def autosize_columns(ws, max_width: int = 36) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def write_output(results: List[Dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Processed curves"
    meta = wb.create_sheet("Processing summary")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write processed data side by side.
    for i, res in enumerate(results):
        col_x = 2 * i + 1
        col_y = 2 * i + 2
        name = res["file_stem"]
        x_type = res["info"]["x_type"]
        x_header = f"{name}_x_{x_type}"
        y_header = f"{name}_end_reaction_kN"
        ws.cell(row=1, column=col_x, value=x_header)
        ws.cell(row=1, column=col_y, value=y_header)
        for c in (col_x, col_y):
            ws.cell(row=1, column=c).font = bold
            ws.cell(row=1, column=c).fill = header_fill
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
            ws.cell(row=1, column=c).border = border

        for r, (xv, yv) in enumerate(zip(res["x"], res["y"]), start=2):
            ws.cell(row=r, column=col_x, value=float(xv))
            ws.cell(row=r, column=col_y, value=float(yv))

    ws.freeze_panes = "A2"
    autosize_columns(ws)

    # Write processing summary.
    summary_headers = [
        "file",
        "data_start_excel_row",
        "x_type",
        "raw_points",
        "processed_points",
        "peak_y_kN_raw_abs",
        "peak_y_kN_processed",
        "slip_correction_applied",
        "x_intercept_removed",
        "elastic_slope",
        "elastic_fit_r2",
        "elastic_linearisation_applied",
        "elastic_linearised_points",
        "elastic_linear_fit_slope",
        "elastic_linear_fit_r2",
        "elastic_fit_start_idx",
        "elastic_fit_end_idx",
        "detected_descending_drops",
    ]
    for c, h in enumerate(summary_headers, start=1):
        cell = meta.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r, res in enumerate(results, start=2):
        info = res["info"]
        values = [
            res["file"],
            res["data_start_row"],
            info.get("x_type"),
            info.get("raw_points"),
            info.get("processed_points"),
            info.get("peak_y_kN_raw_abs"),
            info.get("peak_y_kN_processed"),
            info.get("slip_correction_applied"),
            info.get("x_intercept_removed"),
            info.get("elastic_slope"),
            info.get("elastic_fit_r2"),
            info.get("elastic_linearisation_applied"),
            info.get("elastic_linearised_points"),
            info.get("elastic_linear_fit_slope"),
            info.get("elastic_linear_fit_r2"),
            info.get("elastic_fit_start_idx"),
            info.get("elastic_fit_end_idx"),
            info.get("detected_descending_drops"),
        ]
        for c, v in enumerate(values, start=1):
            meta.cell(row=r, column=c, value=v)

    meta.freeze_panes = "A2"
    autosize_columns(meta)

    wb.save(output_path)


def main() -> None:
    folder = Path(__file__).resolve().parent
    xlsx_files = sorted(
        p for p in folder.glob("*.xlsx") if p.is_file() and not is_ignored_file(p)
    )

    if not xlsx_files:
        raise FileNotFoundError("No RAW .xlsx files found in the same folder as this script.")

    results = []
    print(f"Found {len(xlsx_files)} RAW .xlsx file(s).")

    for path in xlsx_files:
        print(f"Processing: {path.name}")
        try:
            x_raw, y_raw, start_row = read_raw_xy(path, X_COL, Y_COL, SHEET)
            x_out, y_out, info = process_curve(x_raw, y_raw)
            results.append(
                {
                    "file": path.name,
                    "file_stem": path.stem,
                    "data_start_row": start_row,
                    "x": x_out,
                    "y": y_out,
                    "info": info,
                }
            )
            print(
                f"  OK | data starts row {start_row} | {info['x_type']} | "
                f"{info['raw_points']} -> {info['processed_points']} points | "
                f"peak = {info['peak_y_kN_processed']:.3g} kN"
            )
        except Exception as exc:
            print(f"  FAILED: {exc}")

    if not results:
        raise RuntimeError("No files were processed successfully. Check X_COL/Y_COL and file format.")

    output_path = folder / OUTPUT_XLSX
    write_output(results, output_path)
    print(f"\nDone. Output saved to: {output_path.name}")


if __name__ == "__main__":
    main()
